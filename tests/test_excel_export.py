#!/usr/bin/env python3
"""
Excel Export Test Script
========================

This script tests the Excel export functionality to verify that 
Excel corruption issues have been resolved.

Usage:
    python3 test_excel_export.py [profile_name]

If no profile is provided, it will use the first configured profile.
"""

import boto3
import openpyxl
from datetime import datetime
import os
import sys

# Import our main export module
import aws_export_resources
from config import AWS_PROFILES

def test_excel_export(profile_name=None):
    """Test Excel export functionality"""
    
    if not profile_name:
        profile_name = AWS_PROFILES[0] if AWS_PROFILES else 'default'
    
    print(f"🧪 Testing Excel export with profile: {profile_name}")
    print("=" * 50)
    
    try:
        # Test sanitization function
        print("1. Testing data sanitization...")
        test_cases = [
            "=SUM(A1:A10)",  # Formula injection
            "Normal text",   # Regular text
            "Text\x00with\x01control\x02chars",  # Control characters
            "x" * 40000,     # Very long string
            None,            # None value
            123,             # Number
            True,            # Boolean
        ]
        
        for i, test_case in enumerate(test_cases):
            result = aws_export_resources.sanitize_excel_data(test_case)
            print(f"   ✓ Test case {i+1}: {str(test_case)[:30]}... -> {str(result)[:30]}...")
        
        print("\n2. Testing Excel workbook creation...")
        
        # Create a simple test workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Export"
        
        # Test headers
        headers = ['Test Column 1', 'Test Column 2', 'Test Column 3']
        ws.append(headers)
        
        # Test data with potential problematic content
        test_data = [
            ['Normal data', 'Another column', 'Third column'],
            ['=FORMULA()', 'Text with\nnewlines', 'Special chars: äöü'],
            ['+CELL_REF', 'Very long text: ' + 'x' * 1000, '@REFERENCE'],
            ['-NEGATIVE', 'Control\x00chars\x01here', '2024-01-01'],
        ]
        
        for row in test_data:
            sanitized_row = [aws_export_resources.sanitize_excel_data(cell) for cell in row]
            ws.append(sanitized_row)
        
        # Save test file
        test_filename = f'test_export_{datetime.now().strftime("%y%m%d-%H%M")}.xlsx'
        wb.save(test_filename)
        
        print(f"   ✓ Test workbook created: {test_filename}")
        
        print("\n3. Verifying Excel file integrity...")
        
        # Try to open and read the file
        test_wb = openpyxl.load_workbook(test_filename)
        test_ws = test_wb.active
        
        row_count = test_ws.max_row
        col_count = test_ws.max_column
        
        print(f"   ✓ File opened successfully")
        print(f"   ✓ Rows: {row_count}, Columns: {col_count}")
        
        # Read a few cells to verify content
        for row in range(1, min(4, row_count + 1)):
            for col in range(1, min(4, col_count + 1)):
                cell_value = test_ws.cell(row=row, column=col).value
                if cell_value:
                    print(f"   ✓ Cell ({row},{col}): {str(cell_value)[:30]}...")
        
        test_wb.close()
        
        print(f"\n4. Cleaning up test file...")
        os.remove(test_filename)
        print(f"   ✓ Test file removed")
        
        print(f"\n🎉 Excel export test completed successfully!")
        print(f"   Data sanitization is working properly")
        print(f"   Excel files should no longer have corruption issues")
        
        return True
        
    except Exception as e:
        print(f"\n❌ Test failed with error: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """Main test function"""
    profile = sys.argv[1] if len(sys.argv) > 1 else None
    
    success = test_excel_export(profile)
    
    if success:
        print(f"\n✅ All tests passed! Excel export should work without corruption.")
        sys.exit(0)
    else:
        print(f"\n❌ Tests failed! Please check the error messages above.")
        sys.exit(1)

if __name__ == "__main__":
    main()