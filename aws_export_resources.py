import boto3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from botocore.exceptions import ClientError, ProfileNotFound

# ==================== CONFIGURATION ====================
# List your AWS profile names here
AWS_PROFILES = [
    'production',
    'staging',
    'dev'
]

# You can also pass profiles as command line arguments
# python script.py profile1 profile2 profile3
# ======================================================

def get_account_info(session):
    """Get AWS account ID and alias"""
    try:
        sts = session.client('sts')
        account_id = sts.get_caller_identity()['Account']
        
        # Try to get account alias
        iam = session.client('iam')
        try:
            aliases = iam.list_account_aliases()
            account_alias = aliases['AccountAliases'][0] if aliases['AccountAliases'] else f'account-{account_id}'
        except:
            account_alias = f'account-{account_id}'
        
        return account_id, account_alias
    except Exception as e:
        print(f"Error getting account info: {e}")
        return 'unknown', 'unknown'

def get_subnet_details(ec2_client, subnet_id):
    """Get subnet name and CIDR block"""
    try:
        response = ec2_client.describe_subnets(SubnetIds=[subnet_id])
        subnet = response['Subnets'][0]
        subnet_name = next((tag['Value'] for tag in subnet.get('Tags', []) if tag['Key'] == 'Name'), 'N/A')
        return {
            'name': subnet_name,
            'cidr': subnet['CidrBlock'],
            'az': subnet['AvailabilityZone']
        }
    except Exception as e:
        return {'name': 'N/A', 'cidr': 'N/A', 'az': 'N/A'}

def get_vpc_details(ec2_client, vpc_id):
    """Get VPC name and CIDR block"""
    try:
        response = ec2_client.describe_vpcs(VpcIds=[vpc_id])
        vpc = response['Vpcs'][0]
        vpc_name = next((tag['Value'] for tag in vpc.get('Tags', []) if tag['Key'] == 'Name'), 'N/A')
        return {
            'name': vpc_name,
            'cidr': vpc['CidrBlock']
        }
    except Exception as e:
        return {'name': 'N/A', 'cidr': 'N/A'}

def apply_header_style(ws, header_font, header_fill, header_alignment):
    """Apply styling to header row"""
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

def export_ec2_instances(ws, ec2, header_font, header_fill, header_alignment):
    """Export EC2 instances with EBS details"""
    print("  - Exporting EC2 instances...")
    
    headers = [
        'Instance ID', 'Name', 'Type', 'State', 
        'Private IP', 'Public IP', 'Launch Time',
        'Root Device Type', 'Root Device Name', 'Root Volume Size (GB)',
        'Total EBS Volumes', 'Total EBS Size (GB)', 'EBS Volume IDs',
        'EBS Volume Types', 'EBS IOPS', 'EBS Encrypted',
        'VPC ID', 'VPC Name', 'VPC CIDR',
        'Subnet ID', 'Subnet Name', 'Subnet CIDR', 'Availability Zone'
    ]
    ws.append(headers)
    
    try:
        instances = ec2.describe_instances()
        for reservation in instances['Reservations']:
            for instance in reservation['Instances']:
                name = next((tag['Value'] for tag in instance.get('Tags', []) if tag['Key'] == 'Name'), 'N/A')
                vpc_id = instance.get('VpcId', 'N/A')
                subnet_id = instance.get('SubnetId', 'N/A')
                
                vpc_info = get_vpc_details(ec2, vpc_id) if vpc_id != 'N/A' else {'name': 'N/A', 'cidr': 'N/A'}
                subnet_info = get_subnet_details(ec2, subnet_id) if subnet_id != 'N/A' else {'name': 'N/A', 'cidr': 'N/A', 'az': 'N/A'}
                
                # Get EBS volumes information
                block_devices = instance.get('BlockDeviceMappings', [])
                volume_ids = []
                volume_types = []
                volume_sizes = []
                volume_iops = []
                encrypted_statuses = []
                root_volume_size = 'N/A'
                
                for block_device in block_devices:
                    if 'Ebs' in block_device:
                        volume_id = block_device['Ebs'].get('VolumeId')
                        if volume_id:
                            volume_ids.append(volume_id)
                            
                            try:
                                volume_info = ec2.describe_volumes(VolumeIds=[volume_id])['Volumes'][0]
                                volume_types.append(volume_info.get('VolumeType', 'N/A'))
                                size = volume_info.get('Size', 0)
                                volume_sizes.append(size)
                                volume_iops.append(str(volume_info.get('Iops', 'N/A')))
                                encrypted_statuses.append('Yes' if volume_info.get('Encrypted', False) else 'No')
                                
                                if block_device['DeviceName'] == instance.get('RootDeviceName'):
                                    root_volume_size = size
                            except:
                                volume_types.append('N/A')
                                volume_sizes.append(0)
                                volume_iops.append('N/A')
                                encrypted_statuses.append('N/A')
                
                total_ebs_size = sum(volume_sizes)
                
                ws.append([
                    instance['InstanceId'],
                    name,
                    instance['InstanceType'],
                    instance['State']['Name'],
                    instance.get('PrivateIpAddress', 'N/A'),
                    instance.get('PublicIpAddress', 'N/A'),
                    instance['LaunchTime'].strftime('%Y-%m-%d %H:%M:%S'),
                    instance.get('RootDeviceType', 'N/A'),
                    instance.get('RootDeviceName', 'N/A'),
                    root_volume_size,
                    len(volume_ids),
                    total_ebs_size,
                    ', '.join(volume_ids) if volume_ids else 'N/A',
                    ', '.join(volume_types) if volume_types else 'N/A',
                    ', '.join(volume_iops) if volume_iops else 'N/A',
                    ', '.join(encrypted_statuses) if encrypted_statuses else 'N/A',
                    vpc_id,
                    vpc_info['name'],
                    vpc_info['cidr'],
                    subnet_id,
                    subnet_info['name'],
                    subnet_info['cidr'],
                    subnet_info['az']
                ])
    except Exception as e:
        print(f"    Error: {e}")
    
    apply_header_style(ws, header_font, header_fill, header_alignment)

def export_rds_instances(ws, rds, ec2, header_font, header_fill, header_alignment):
    """Export RDS instances"""
    print("  - Exporting RDS instances...")
    
    headers = [
        'DB Identifier', 'Engine', 'Engine Version', 
        'Instance Class', 'Status', 'Endpoint', 'Port', 'Storage (GB)',
        'Multi-AZ', 'VPC ID', 'VPC Name', 'VPC CIDR',
        'Subnet Group', 'Subnets', 'Availability Zone'
    ]
    ws.append(headers)
    
    try:
        db_instances = rds.describe_db_instances()
        for db in db_instances['DBInstances']:
            endpoint = db.get('Endpoint', {}).get('Address', 'N/A')
            port = db.get('Endpoint', {}).get('Port', 'N/A')
            
            vpc_id = db.get('DBSubnetGroup', {}).get('VpcId', 'N/A')
            vpc_info = get_vpc_details(ec2, vpc_id) if vpc_id != 'N/A' else {'name': 'N/A', 'cidr': 'N/A'}
            
            subnet_group_name = db.get('DBSubnetGroup', {}).get('DBSubnetGroupName', 'N/A')
            subnets = db.get('DBSubnetGroup', {}).get('Subnets', [])
            subnet_list = ', '.join([s['SubnetIdentifier'] for s in subnets])
            
            ws.append([
                db['DBInstanceIdentifier'],
                db['Engine'],
                db['EngineVersion'],
                db['DBInstanceClass'],
                db['DBInstanceStatus'],
                endpoint,
                port,
                db['AllocatedStorage'],
                'Yes' if db['MultiAZ'] else 'No',
                vpc_id,
                vpc_info['name'],
                vpc_info['cidr'],
                subnet_group_name,
                subnet_list,
                db['AvailabilityZone']
            ])
    except Exception as e:
        print(f"    Error: {e}")
    
    apply_header_style(ws, header_font, header_fill, header_alignment)

def export_rds_clusters(ws, rds, ec2, header_font, header_fill, header_alignment):
    """Export RDS Aurora Clusters"""
    print("  - Exporting RDS Clusters (Aurora)...")
    
    headers = [
        'Cluster Identifier', 'Engine', 'Engine Version', 'Engine Mode',
        'Status', 'Cluster Endpoint', 'Reader Endpoint', 'Port',
        'Database Name', 'Master Username', 'Multi-AZ',
        'Cluster Members', 'Storage Encrypted', 'Backup Retention (Days)',
        'VPC ID', 'VPC Name', 'VPC CIDR',
        'Subnet Group', 'Availability Zones', 'Storage Type'
    ]
    ws.append(headers)
    
    try:
        db_clusters = rds.describe_db_clusters()
        for cluster in db_clusters['DBClusters']:
            # Get endpoints
            cluster_endpoint = cluster.get('Endpoint', 'N/A')
            reader_endpoint = cluster.get('ReaderEndpoint', 'N/A')
            
            # Get VPC info
            vpc_id = 'N/A'
            vpc_info = {'name': 'N/A', 'cidr': 'N/A'}
            
            # Try to get VPC from cluster members
            if cluster.get('DBClusterMembers'):
                try:
                    member_id = cluster['DBClusterMembers'][0]['DBInstanceIdentifier']
                    member_info = rds.describe_db_instances(DBInstanceIdentifier=member_id)
                    if member_info['DBInstances']:
                        vpc_id = member_info['DBInstances'][0].get('DBSubnetGroup', {}).get('VpcId', 'N/A')
                        if vpc_id != 'N/A':
                            vpc_info = get_vpc_details(ec2, vpc_id)
                except:
                    pass
            
            # Get subnet group
            subnet_group_name = cluster.get('DBSubnetGroup', 'N/A')
            
            # Get availability zones
            availability_zones = ', '.join(cluster.get('AvailabilityZones', []))
            
            # Get cluster members count
            cluster_members = len(cluster.get('DBClusterMembers', []))
            member_list = ', '.join([m['DBInstanceIdentifier'] for m in cluster.get('DBClusterMembers', [])])
            
            ws.append([
                cluster['DBClusterIdentifier'],
                cluster['Engine'],
                cluster['EngineVersion'],
                cluster.get('EngineMode', 'provisioned'),
                cluster['Status'],
                cluster_endpoint,
                reader_endpoint,
                cluster.get('Port', 'N/A'),
                cluster.get('DatabaseName', 'N/A'),
                cluster.get('MasterUsername', 'N/A'),
                'Yes' if cluster.get('MultiAZ', False) else 'No',
                f"{cluster_members} ({member_list})" if member_list else str(cluster_members),
                'Yes' if cluster.get('StorageEncrypted', False) else 'No',
                cluster.get('BackupRetentionPeriod', 0),
                vpc_id,
                vpc_info['name'],
                vpc_info['cidr'],
                subnet_group_name,
                availability_zones if availability_zones else 'N/A',
                cluster.get('StorageType', 'aurora')
            ])
    except Exception as e:
        print(f"    Error: {e}")
    
    apply_header_style(ws, header_font, header_fill, header_alignment)

def export_lambda_functions(ws, lambda_client, ec2, header_font, header_fill, header_alignment):
    """Export Lambda functions"""
    print("  - Exporting Lambda functions...")
    
    headers = [
        'Function Name', 'Runtime', 'Memory (MB)', 
        'Timeout (sec)', 'Last Modified', 'Handler',
        'VPC ID', 'VPC Name', 'VPC CIDR',
        'Subnet IDs', 'Subnet Names', 'Security Groups'
    ]
    ws.append(headers)
    
    try:
        functions = lambda_client.list_functions()
        for func in functions['Functions']:
            vpc_config = func.get('VpcConfig', {})
            vpc_id = vpc_config.get('VpcId', 'No VPC')
            subnet_ids = vpc_config.get('SubnetIds', [])
            security_groups = vpc_config.get('SecurityGroupIds', [])
            
            if vpc_id != 'No VPC':
                vpc_info = get_vpc_details(ec2, vpc_id)
                subnet_names = []
                for subnet_id in subnet_ids:
                    subnet_info = get_subnet_details(ec2, subnet_id)
                    subnet_names.append(subnet_info['name'])
            else:
                vpc_info = {'name': 'N/A', 'cidr': 'N/A'}
                subnet_names = []
            
            ws.append([
                func['FunctionName'],
                func.get('Runtime', 'N/A'),
                func['MemorySize'],
                func['Timeout'],
                func['LastModified'],
                func['Handler'],
                vpc_id,
                vpc_info['name'],
                vpc_info['cidr'],
                ', '.join(subnet_ids) if subnet_ids else 'N/A',
                ', '.join(subnet_names) if subnet_names else 'N/A',
                ', '.join(security_groups) if security_groups else 'N/A'
            ])
    except Exception as e:
        print(f"    Error: {e}")
    
    apply_header_style(ws, header_font, header_fill, header_alignment)

def export_efs_filesystems(ws, efs, header_font, header_fill, header_alignment):
    """Export EFS file systems"""
    print("  - Exporting EFS file systems...")
    
    headers = [
        'File System ID', 'Name', 'Creation Time', 'Life Cycle State',
        'Performance Mode', 'Throughput Mode', 'Encrypted', 'Size (GB)',
        'Mount Targets', 'VPC IDs', 'Subnet IDs', 'Availability Zones'
    ]
    ws.append(headers)
    
    try:
        file_systems = efs.describe_file_systems()
        for fs in file_systems['FileSystems']:
            fs_name = fs.get('Name', 'N/A')
            
            try:
                mount_targets = efs.describe_mount_targets(FileSystemId=fs['FileSystemId'])
                mt_count = len(mount_targets['MountTargets'])
                vpc_ids = list(set([mt.get('VpcId', 'N/A') for mt in mount_targets['MountTargets']]))
                subnet_ids = [mt.get('SubnetId', 'N/A') for mt in mount_targets['MountTargets']]
                azs = [mt.get('AvailabilityZoneName', 'N/A') for mt in mount_targets['MountTargets']]
            except:
                mt_count = 0
                vpc_ids = ['N/A']
                subnet_ids = ['N/A']
                azs = ['N/A']
            
            size_gb = fs.get('SizeInBytes', {}).get('Value', 0) / (1024**3)
            
            ws.append([
                fs['FileSystemId'],
                fs_name,
                fs['CreationTime'].strftime('%Y-%m-%d %H:%M:%S'),
                fs['LifeCycleState'],
                fs.get('PerformanceMode', 'N/A'),
                fs.get('ThroughputMode', 'N/A'),
                'Yes' if fs.get('Encrypted', False) else 'No',
                f"{size_gb:.2f}",
                mt_count,
                ', '.join(vpc_ids),
                ', '.join(subnet_ids),
                ', '.join(azs)
            ])
    except Exception as e:
        print(f"    Error: {e}")
    
    apply_header_style(ws, header_font, header_fill, header_alignment)

def export_ecs_services(ws, ecs, ec2, header_font, header_fill, header_alignment):
    """Export ECS services"""
    print("  - Exporting ECS services...")
    
    headers = [
        'Cluster Name', 'Service Name', 'Status', 'Desired Count',
        'Running Count', 'Launch Type', 'Task Definition',
        'Storage Type', 'Volume Size (GB)', 'Volume Type', 'Volume IOPS',
        'VPC ID', 'Subnet IDs', 'Security Groups', 'Load Balancers'
    ]
    ws.append(headers)
    
    try:
        clusters = ecs.list_clusters()
        for cluster_arn in clusters.get('clusterArns', []):
            cluster_name = cluster_arn.split('/')[-1]
            
            services = ecs.list_services(cluster=cluster_arn)
            if services.get('serviceArns'):
                service_details = ecs.describe_services(
                    cluster=cluster_arn,
                    services=services['serviceArns']
                )
                
                for service in service_details.get('services', []):
                    network_config = service.get('networkConfiguration', {}).get('awsvpcConfiguration', {})
                    vpc_id = 'N/A'
                    subnets = network_config.get('subnets', [])
                    security_groups = network_config.get('securityGroups', [])
                    
                    if subnets:
                        try:
                            subnet_response = ec2.describe_subnets(SubnetIds=[subnets[0]])
                            vpc_id = subnet_response['Subnets'][0]['VpcId']
                        except:
                            pass
                    
                    load_balancers = ', '.join([lb.get('targetGroupArn', '').split('/')[-1] 
                                                for lb in service.get('loadBalancers', [])])
                    
                    task_def_arn = service['taskDefinition']
                    storage_type = 'N/A'
                    volume_size = 'N/A'
                    volume_type = 'N/A'
                    volume_iops = 'N/A'
                    
                    try:
                        task_def = ecs.describe_task_definition(taskDefinition=task_def_arn)['taskDefinition']
                        
                        ephemeral_storage = task_def.get('ephemeralStorage', {})
                        if ephemeral_storage:
                            storage_type = 'Ephemeral (Fargate)'
                            volume_size = ephemeral_storage.get('sizeInGiB', 20)
                        
                        volumes = task_def.get('volumes', [])
                        ebs_volumes = []
                        for volume in volumes:
                            if 'efsVolumeConfiguration' in volume:
                                storage_type = 'EFS'
                            elif 'dockerVolumeConfiguration' in volume:
                                docker_config = volume['dockerVolumeConfiguration']
                                driver_opts = docker_config.get('driverOpts', {})
                                if 'size' in driver_opts:
                                    ebs_volumes.append({
                                        'size': driver_opts.get('size', 'N/A'),
                                        'type': driver_opts.get('type', 'gp3'),
                                        'iops': driver_opts.get('iops', 'N/A')
                                    })
                        
                        if ebs_volumes:
                            storage_type = 'EBS Volume'
                            volume_size = ', '.join([str(v['size']) for v in ebs_volumes])
                            volume_type = ', '.join([str(v['type']) for v in ebs_volumes])
                            volume_iops = ', '.join([str(v['iops']) for v in ebs_volumes])
                        elif storage_type == 'N/A' and service.get('launchType') == 'FARGATE':
                            storage_type = 'Ephemeral (Fargate)'
                            volume_size = 20
                        elif storage_type == 'N/A':
                            storage_type = 'Host/Container'
                            
                    except Exception as e:
                        pass
                    
                    ws.append([
                        cluster_name,
                        service['serviceName'],
                        service['status'],
                        service['desiredCount'],
                        service['runningCount'],
                        service.get('launchType', 'N/A'),
                        service['taskDefinition'].split('/')[-1],
                        storage_type,
                        volume_size,
                        volume_type,
                        volume_iops,
                        vpc_id,
                        ', '.join(subnets) if subnets else 'N/A',
                        ', '.join(security_groups) if security_groups else 'N/A',
                        load_balancers if load_balancers else 'N/A'
                    ])
    except Exception as e:
        print(f"    Error: {e}")
    
    apply_header_style(ws, header_font, header_fill, header_alignment)

def export_eks_clusters(ws, eks, ec2, header_font, header_fill, header_alignment):
    """Export EKS clusters"""
    print("  - Exporting EKS clusters...")
    
    headers = [
        'Cluster Name', 'Version', 'Status', 'Endpoint',
        'Created At', 'VPC ID', 'VPC Name', 'VPC CIDR',
        'Subnet IDs', 'Security Group IDs', 'Role ARN'
    ]
    ws.append(headers)
    
    try:
        clusters = eks.list_clusters()
        for cluster_name in clusters.get('clusters', []):
            cluster = eks.describe_cluster(name=cluster_name)['cluster']
            
            vpc_id = cluster.get('resourcesVpcConfig', {}).get('vpcId', 'N/A')
            vpc_info = get_vpc_details(ec2, vpc_id) if vpc_id != 'N/A' else {'name': 'N/A', 'cidr': 'N/A'}
            
            subnet_ids = cluster.get('resourcesVpcConfig', {}).get('subnetIds', [])
            security_group_ids = cluster.get('resourcesVpcConfig', {}).get('securityGroupIds', [])
            
            ws.append([
                cluster['name'],
                cluster.get('version', 'N/A'),
                cluster['status'],
                cluster.get('endpoint', 'N/A'),
                cluster['createdAt'].strftime('%Y-%m-%d %H:%M:%S'),
                vpc_id,
                vpc_info['name'],
                vpc_info['cidr'],
                ', '.join(subnet_ids) if subnet_ids else 'N/A',
                ', '.join(security_group_ids) if security_group_ids else 'N/A',
                cluster.get('roleArn', 'N/A')
            ])
    except Exception as e:
        print(f"    Error: {e}")
    
    apply_header_style(ws, header_font, header_fill, header_alignment)

def export_elasticache_clusters(ws, elasticache, ec2, header_font, header_fill, header_alignment):
    """Export ElastiCache clusters"""
    print("  - Exporting ElastiCache clusters...")
    
    headers = [
        'Cluster ID', 'Engine', 'Engine Version', 'Node Type',
        'Status', 'Num Cache Nodes', 'Preferred AZ',
        'VPC ID', 'VPC Name', 'VPC CIDR', 'Subnet Group',
        'Security Groups', 'Endpoint'
    ]
    ws.append(headers)
    
    try:
        redis_clusters = elasticache.describe_replication_groups()
        for cluster in redis_clusters.get('ReplicationGroups', []):
            node_groups = cluster.get('NodeGroups', [{}])[0]
            endpoint = node_groups.get('PrimaryEndpoint', {}).get('Address', 'N/A')
            
            cache_cluster_id = cluster.get('MemberClusters', [''])[0]
            vpc_id = 'N/A'
            vpc_info = {'name': 'N/A', 'cidr': 'N/A'}
            subnet_group_name = 'N/A'
            security_groups = []
            
            if cache_cluster_id:
                try:
                    cache_cluster = elasticache.describe_cache_clusters(
                        CacheClusterId=cache_cluster_id
                    )['CacheClusters'][0]
                    subnet_group_name = cache_cluster.get('CacheSubnetGroupName', 'N/A')
                    security_groups = [sg['SecurityGroupId'] for sg in cache_cluster.get('SecurityGroups', [])]
                    
                    if subnet_group_name != 'N/A':
                        subnet_group = elasticache.describe_cache_subnet_groups(
                            CacheSubnetGroupName=subnet_group_name
                        )['CacheSubnetGroups'][0]
                        vpc_id = subnet_group.get('VpcId', 'N/A')
                        vpc_info = get_vpc_details(ec2, vpc_id) if vpc_id != 'N/A' else {'name': 'N/A', 'cidr': 'N/A'}
                except:
                    pass
            
            ws.append([
                cluster['ReplicationGroupId'],
                'Redis',
                cluster.get('CacheNodeType', 'N/A'),
                cluster.get('CacheNodeType', 'N/A'),
                cluster['Status'],
                len(cluster.get('MemberClusters', [])),
                node_groups.get('PrimaryEndpoint', {}).get('AvailabilityZone', 'N/A'),
                vpc_id,
                vpc_info['name'],
                vpc_info['cidr'],
                subnet_group_name,
                ', '.join(security_groups) if security_groups else 'N/A',
                endpoint
            ])
        
        memcached_clusters = elasticache.describe_cache_clusters()
        for cluster in memcached_clusters.get('CacheClusters', []):
            if cluster.get('Engine') == 'memcached':
                subnet_group_name = cluster.get('CacheSubnetGroupName', 'N/A')
                vpc_id = 'N/A'
                vpc_info = {'name': 'N/A', 'cidr': 'N/A'}
                
                if subnet_group_name != 'N/A':
                    try:
                        subnet_group = elasticache.describe_cache_subnet_groups(
                            CacheSubnetGroupName=subnet_group_name
                        )['CacheSubnetGroups'][0]
                        vpc_id = subnet_group.get('VpcId', 'N/A')
                        vpc_info = get_vpc_details(ec2, vpc_id) if vpc_id != 'N/A' else {'name': 'N/A', 'cidr': 'N/A'}
                    except:
                        pass
                
                security_groups = [sg['SecurityGroupId'] for sg in cluster.get('SecurityGroups', [])]
                endpoint = cluster.get('ConfigurationEndpoint', {}).get('Address', 'N/A')
                
                ws.append([
                    cluster['CacheClusterId'],
                    cluster['Engine'],
                    cluster.get('EngineVersion', 'N/A'),
                    cluster.get('CacheNodeType', 'N/A'),
                    cluster['CacheClusterStatus'],
                    cluster.get('NumCacheNodes', 0),
                    cluster.get('PreferredAvailabilityZone', 'N/A'),
                    vpc_id,
                    vpc_info['name'],
                    vpc_info['cidr'],
                    subnet_group_name,
                    ', '.join(security_groups) if security_groups else 'N/A',
                    endpoint
                ])
    except Exception as e:
        print(f"    Error: {e}")
    
    apply_header_style(ws, header_font, header_fill, header_alignment)

def export_mq_brokers(ws, mq, ec2, header_font, header_fill, header_alignment):
    """Export Amazon MQ brokers"""
    print("  - Exporting Amazon MQ brokers...")
    
    headers = [
        'Broker Name', 'Broker ID', 'Engine Type', 'Engine Version',
        'Deployment Mode', 'Instance Type', 'Status',
        'VPC ID', 'VPC Name', 'VPC CIDR',
        'Subnet IDs', 'Security Groups', 'Endpoints'
    ]
    ws.append(headers)
    
    try:
        brokers = mq.list_brokers()
        for broker_summary in brokers.get('BrokerSummaries', []):
            broker = mq.describe_broker(BrokerId=broker_summary['BrokerId'])
            
            vpc_id = 'N/A'
            vpc_info = {'name': 'N/A', 'cidr': 'N/A'}
            subnet_ids = broker.get('SubnetIds', [])
            
            if subnet_ids:
                try:
                    subnet_response = ec2.describe_subnets(SubnetIds=[subnet_ids[0]])
                    vpc_id = subnet_response['Subnets'][0]['VpcId']
                    vpc_info = get_vpc_details(ec2, vpc_id)
                except:
                    pass
            
            security_groups = broker.get('SecurityGroups', [])
            
            endpoints = []
            for instance in broker.get('BrokerInstances', []):
                if instance.get('Endpoints'):
                    endpoints.extend(instance['Endpoints'])
            
            ws.append([
                broker['BrokerName'],
                broker['BrokerId'],
                broker['EngineType'],
                broker.get('EngineVersion', 'N/A'),
                broker['DeploymentMode'],
                broker['HostInstanceType'],
                broker['BrokerState'],
                vpc_id,
                vpc_info['name'],
                vpc_info['cidr'],
                ', '.join(subnet_ids) if subnet_ids else 'N/A',
                ', '.join(security_groups) if security_groups else 'N/A',
                '\n'.join(endpoints[:3]) if endpoints else 'N/A'
            ])
    except Exception as e:
        print(f"    Error: {e}")
    
    apply_header_style(ws, header_font, header_fill, header_alignment)

def export_load_balancers(ws, elbv2, ec2, header_font, header_fill, header_alignment):
    """Export Elastic Load Balancers (ALB, NLB, GLB)"""
    print("  - Exporting Load Balancers...")
    
    headers = [
        'Load Balancer Name', 'Type', 'Scheme', 'State',
        'DNS Name', 'Created Time',
        'VPC ID', 'VPC Name', 'VPC CIDR',
        'Availability Zones', 'Subnet IDs', 'Security Groups',
        'IP Address Type'
    ]
    ws.append(headers)
    
    try:
        load_balancers = elbv2.describe_load_balancers()
        
        for lb in load_balancers['LoadBalancers']:
            vpc_id = lb.get('VpcId', 'N/A')
            vpc_info = get_vpc_details(ec2, vpc_id) if vpc_id != 'N/A' else {'name': 'N/A', 'cidr': 'N/A'}
            
            # Get availability zones and subnets
            azs = [az['ZoneName'] for az in lb.get('AvailabilityZones', [])]
            subnet_ids = [az['SubnetId'] for az in lb.get('AvailabilityZones', [])]
            
            security_groups = lb.get('SecurityGroups', [])
            
            ws.append([
                lb['LoadBalancerName'],
                lb['Type'],
                lb['Scheme'],
                lb['State']['Code'],
                lb['DNSName'],
                lb['CreatedTime'].strftime('%Y-%m-%d %H:%M:%S'),
                vpc_id,
                vpc_info['name'],
                vpc_info['cidr'],
                ', '.join(azs) if azs else 'N/A',
                ', '.join(subnet_ids) if subnet_ids else 'N/A',
                ', '.join(security_groups) if security_groups else 'N/A',
                lb.get('IpAddressType', 'N/A')
            ])
    except Exception as e:
        print(f"    Error: {e}")
    
    apply_header_style(ws, header_font, header_fill, header_alignment)

def export_personalize(ws, personalize, header_font, header_fill, header_alignment):
    """Export AWS Personalize resources"""
    print("  - Exporting AWS Personalize...")
    
    headers = [
        'Campaign Name', 'Campaign ARN', 'Status',
        'Solution Version ARN', 'Min Provisioned TPS',
        'Created Time', 'Last Updated'
    ]
    ws.append(headers)
    
    try:
        campaigns = personalize.list_campaigns()
        
        for campaign_summary in campaigns.get('campaigns', []):
            campaign_arn = campaign_summary['campaignArn']
            
            try:
                campaign = personalize.describe_campaign(campaignArn=campaign_arn)['campaign']
                
                ws.append([
                    campaign['name'],
                    campaign['campaignArn'],
                    campaign['status'],
                    campaign.get('solutionVersionArn', 'N/A'),
                    campaign.get('minProvisionedTPS', 'N/A'),
                    campaign['creationDateTime'].strftime('%Y-%m-%d %H:%M:%S'),
                    campaign['lastUpdatedDateTime'].strftime('%Y-%m-%d %H:%M:%S')
                ])
            except:
                pass
    except Exception as e:
        print(f"    Error: {e}")
    
    apply_header_style(ws, header_font, header_fill, header_alignment)

def export_cloudwatch_alarms(ws, cloudwatch, header_font, header_fill, header_alignment):
    """Export CloudWatch Alarms"""
    print("  - Exporting CloudWatch Alarms...")
    
    headers = [
        'Alarm Name', 'State', 'Metric Name', 'Namespace',
        'Statistic', 'Period (sec)', 'Threshold', 'Comparison Operator',
        'Evaluation Periods', 'Actions Enabled', 'Alarm Actions',
        'Dimensions'
    ]
    ws.append(headers)
    
    try:
        alarms = cloudwatch.describe_alarms()
        
        for alarm in alarms['MetricAlarms']:
            dimensions = ', '.join([f"{d['Name']}={d['Value']}" for d in alarm.get('Dimensions', [])])
            alarm_actions = ', '.join(alarm.get('AlarmActions', []))
            
            ws.append([
                alarm['AlarmName'],
                alarm['StateValue'],
                alarm.get('MetricName', 'N/A'),
                alarm.get('Namespace', 'N/A'),
                alarm.get('Statistic', 'N/A'),
                alarm.get('Period', 'N/A'),
                alarm.get('Threshold', 'N/A'),
                alarm.get('ComparisonOperator', 'N/A'),
                alarm.get('EvaluationPeriods', 'N/A'),
                'Yes' if alarm.get('ActionsEnabled', False) else 'No',
                alarm_actions if alarm_actions else 'N/A',
                dimensions if dimensions else 'N/A'
            ])
    except Exception as e:
        print(f"    Error: {e}")
    
    apply_header_style(ws, header_font, header_fill, header_alignment)

def export_cloudwatch_log_groups(ws, logs, header_font, header_fill, header_alignment):
    """Export CloudWatch Log Groups"""
    print("  - Exporting CloudWatch Log Groups...")
    
    headers = [
        'Log Group Name', 'Creation Time', 'Retention (Days)',
        'Stored Bytes', 'Metric Filter Count', 'KMS Key ID'
    ]
    ws.append(headers)
    
    try:
        paginator = logs.get_paginator('describe_log_groups')
        
        for page in paginator.paginate():
            for log_group in page['logGroups']:
                stored_bytes = log_group.get('storedBytes', 0)
                stored_mb = stored_bytes / (1024 * 1024)
                
                creation_time = datetime.fromtimestamp(log_group['creationTime'] / 1000).strftime('%Y-%m-%d %H:%M:%S')
                
                ws.append([
                    log_group['logGroupName'],
                    creation_time,
                    log_group.get('retentionInDays', 'Never Expire'),
                    f"{stored_mb:.2f} MB",
                    log_group.get('metricFilterCount', 0),
                    log_group.get('kmsKeyId', 'N/A')
                ])
    except Exception as e:
        print(f"    Error: {e}")
    
    apply_header_style(ws, header_font, header_fill, header_alignment)

def export_dynamodb_tables(ws, dynamodb, header_font, header_fill, header_alignment):
    """Export DynamoDB Tables"""
    print("  - Exporting DynamoDB Tables...")
    
    headers = [
        'Table Name', 'Status', 'Creation Time',
        'Item Count', 'Table Size (Bytes)', 'Read Capacity Units',
        'Write Capacity Units', 'Billing Mode', 'Table Class',
        'Partition Key', 'Sort Key', 'Global Secondary Indexes',
        'Local Secondary Indexes', 'Stream Enabled', 'Point-in-time Recovery',
        'Encryption Type'
    ]
    ws.append(headers)
    
    try:
        tables = dynamodb.list_tables()
        
        for table_name in tables['TableNames']:
            try:
                table = dynamodb.describe_table(TableName=table_name)['Table']
                
                # Get partition and sort keys
                partition_key = 'N/A'
                sort_key = 'N/A'
                for key in table.get('KeySchema', []):
                    if key['KeyType'] == 'HASH':
                        partition_key = key['AttributeName']
                    elif key['KeyType'] == 'RANGE':
                        sort_key = key['AttributeName']
                
                # Get GSI and LSI
                gsi_count = len(table.get('GlobalSecondaryIndexes', []))
                lsi_count = len(table.get('LocalSecondaryIndexes', []))
                
                # Get capacity info
                billing_mode = table.get('BillingModeSummary', {}).get('BillingMode', 
                               'PROVISIONED' if table.get('ProvisionedThroughput') else 'N/A')
                
                read_capacity = table.get('ProvisionedThroughput', {}).get('ReadCapacityUnits', 'On-Demand')
                write_capacity = table.get('ProvisionedThroughput', {}).get('WriteCapacityUnits', 'On-Demand')
                
                # Get additional features
                stream_enabled = 'Yes' if table.get('StreamSpecification', {}).get('StreamEnabled', False) else 'No'
                
                # Point-in-time recovery
                try:
                    pitr = dynamodb.describe_continuous_backups(TableName=table_name)
                    pitr_status = pitr['ContinuousBackupsDescription']['PointInTimeRecoveryDescription']['PointInTimeRecoveryStatus']
                    pitr_enabled = 'Yes' if pitr_status == 'ENABLED' else 'No'
                except:
                    pitr_enabled = 'N/A'
                
                encryption_type = table.get('SSEDescription', {}).get('SSEType', 'N/A')
                
                ws.append([
                    table['TableName'],
                    table['TableStatus'],
                    table['CreationDateTime'].strftime('%Y-%m-%d %H:%M:%S'),
                    table.get('ItemCount', 0),
                    table.get('TableSizeBytes', 0),
                    read_capacity,
                    write_capacity,
                    billing_mode,
                    table.get('TableClassSummary', {}).get('TableClass', 'STANDARD'),
                    partition_key,
                    sort_key,
                    gsi_count,
                    lsi_count,
                    stream_enabled,
                    pitr_enabled,
                    encryption_type
                ])
            except Exception as e:
                print(f"    Error processing table {table_name}: {e}")
    except Exception as e:
        print(f"    Error: {e}")
    
    apply_header_style(ws, header_font, header_fill, header_alignment)

def export_transfer_family(ws, transfer, ec2, header_font, header_fill, header_alignment):
    """Export AWS Transfer Family Servers"""
    print("  - Exporting AWS Transfer Family...")
    
    headers = [
        'Server ID', 'State', 'Endpoint Type', 'Domain',
        'Protocols', 'Identity Provider Type', 'Logging Role',
        'VPC ID', 'VPC Name', 'VPC CIDR', 'Subnet IDs',
        'Security Group IDs', 'User Count'
    ]
    ws.append(headers)
    
    try:
        servers = transfer.list_servers()
        
        for server_summary in servers.get('Servers', []):
            server_id = server_summary['ServerId']
            
            try:
                server = transfer.describe_server(ServerId=server_id)['Server']
                
                # Get VPC details for VPC endpoint type
                vpc_id = 'N/A'
                vpc_info = {'name': 'N/A', 'cidr': 'N/A'}
                subnet_ids = []
                security_groups = []
                
                if server.get('EndpointType') == 'VPC':
                    endpoint_details = server.get('EndpointDetails', {})
                    subnet_ids = endpoint_details.get('SubnetIds', [])
                    security_groups = endpoint_details.get('SecurityGroupIds', [])
                    vpc_id = endpoint_details.get('VpcId', 'N/A')
                    
                    if vpc_id != 'N/A':
                        vpc_info = get_vpc_details(ec2, vpc_id)
                
                # Get user count
                try:
                    users = transfer.list_users(ServerId=server_id)
                    user_count = len(users.get('Users', []))
                except:
                    user_count = 0
                
                protocols = ', '.join(server.get('Protocols', []))
                
                ws.append([
                    server['ServerId'],
                    server['State'],
                    server.get('EndpointType', 'N/A'),
                    server.get('Domain', 'N/A'),
                    protocols,
                    server.get('IdentityProviderType', 'N/A'),
                    server.get('LoggingRole', 'N/A'),
                    vpc_id,
                    vpc_info['name'],
                    vpc_info['cidr'],
                    ', '.join(subnet_ids) if subnet_ids else 'N/A',
                    ', '.join(security_groups) if security_groups else 'N/A',
                    user_count
                ])
            except Exception as e:
                print(f"    Error processing server {server_id}: {e}")
    except Exception as e:
        print(f"    Error: {e}")
    
    apply_header_style(ws, header_font, header_fill, header_alignment)

def export_vpc_summary(ws, ec2, region, header_font, header_fill, header_alignment):
    """Export VPC summary"""
    print("  - Exporting VPC summary...")
    
    headers = ['VPC ID', 'VPC Name', 'CIDR Block', 'State', 'Default VPC', 'Region']
    ws.append(headers)
    
    try:
        vpcs = ec2.describe_vpcs()
        
        for vpc in vpcs['Vpcs']:
            vpc_name = next((tag['Value'] for tag in vpc.get('Tags', []) if tag['Key'] == 'Name'), 'N/A')
            
            ws.append([
                vpc['VpcId'],
                vpc_name,
                vpc['CidrBlock'],
                vpc['State'],
                'Yes' if vpc.get('IsDefault', False) else 'No',
                region
            ])
    except Exception as e:
        print(f"    Error: {e}")
    
    apply_header_style(ws, header_font, header_fill, header_alignment)

def export_s3_buckets(ws, s3, header_font, header_fill, header_alignment):
    """Export S3 Buckets with detailed information"""
    print("  - Exporting S3 Buckets...")
    
    headers = [
        'Bucket Name', 'Creation Date', 'Region',
        'Versioning Status', 'Encryption Type', 'Public Access Block',
        'Total Objects', 'Total Size (GB)', 'Storage Classes',
        'Lifecycle Rules', 'Replication Status', 'Logging Enabled',
        'Website Hosting', 'Tags'
    ]
    ws.append(headers)
    
    try:
        buckets = s3.list_buckets()
        
        for bucket in buckets.get('Buckets', []):
            bucket_name = bucket['Name']
            
            try:
                # Get bucket location
                try:
                    location = s3.get_bucket_location(Bucket=bucket_name)
                    region = location['LocationConstraint'] or 'us-east-1'
                except:
                    region = 'N/A'
                
                # Get versioning status
                try:
                    versioning = s3.get_bucket_versioning(Bucket=bucket_name)
                    versioning_status = versioning.get('Status', 'Disabled')
                except:
                    versioning_status = 'N/A'
                
                # Get encryption
                try:
                    encryption = s3.get_bucket_encryption(Bucket=bucket_name)
                    encryption_type = encryption['ServerSideEncryptionConfiguration']['Rules'][0]['ApplyServerSideEncryptionByDefault']['SSEAlgorithm']
                except:
                    encryption_type = 'None'
                
                # Get public access block
                try:
                    public_access = s3.get_public_access_block(Bucket=bucket_name)
                    config = public_access['PublicAccessBlockConfiguration']
                    if all([config.get('BlockPublicAcls'), config.get('BlockPublicPolicy'),
                           config.get('IgnorePublicAcls'), config.get('RestrictPublicBuckets')]):
                        public_access_status = 'All Blocked'
                    else:
                        public_access_status = 'Partially Blocked'
                except:
                    public_access_status = 'Not Configured'
                
                # Get bucket metrics (object count and size)
                try:
                    cloudwatch = boto3.client('cloudwatch', region_name=region)
                    # Note: This requires CloudWatch metrics to be enabled
                    total_objects = 'N/A'
                    total_size_gb = 'N/A'
                    storage_classes = 'N/A'
                except:
                    total_objects = 'N/A'
                    total_size_gb = 'N/A'
                    storage_classes = 'N/A'
                
                # Get lifecycle rules
                try:
                    lifecycle = s3.get_bucket_lifecycle_configuration(Bucket=bucket_name)
                    lifecycle_count = len(lifecycle.get('Rules', []))
                except:
                    lifecycle_count = 0
                
                # Get replication status
                try:
                    replication = s3.get_bucket_replication(Bucket=bucket_name)
                    replication_status = 'Enabled'
                except:
                    replication_status = 'Disabled'
                
                # Get logging status
                try:
                    logging = s3.get_bucket_logging(Bucket=bucket_name)
                    logging_enabled = 'Yes' if logging.get('LoggingEnabled') else 'No'
                except:
                    logging_enabled = 'No'
                
                # Get website hosting
                try:
                    website = s3.get_bucket_website(Bucket=bucket_name)
                    website_hosting = 'Enabled'
                except:
                    website_hosting = 'Disabled'
                
                # Get tags
                try:
                    tags = s3.get_bucket_tagging(Bucket=bucket_name)
                    tag_list = ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags.get('TagSet', [])])
                except:
                    tag_list = 'N/A'
                
                ws.append([
                    bucket_name,
                    bucket['CreationDate'].strftime('%Y-%m-%d %H:%M:%S'),
                    region,
                    versioning_status,
                    encryption_type,
                    public_access_status,
                    total_objects,
                    total_size_gb,
                    storage_classes,
                    lifecycle_count,
                    replication_status,
                    logging_enabled,
                    website_hosting,
                    tag_list
                ])
            except Exception as e:
                print(f"    Error processing bucket {bucket_name}: {e}")
                
    except Exception as e:
        print(f"    Error: {e}")
    
    apply_header_style(ws, header_font, header_fill, header_alignment)

def export_s3_glacier_vaults(ws, glacier, header_font, header_fill, header_alignment):
    """Export S3 Glacier Vaults"""
    print("  - Exporting S3 Glacier Vaults...")
    
    headers = [
        'Vault Name', 'Vault ARN', 'Creation Date',
        'Last Inventory Date', 'Number of Archives', 'Size (GB)',
        'Notifications Enabled', 'SNS Topic'
    ]
    ws.append(headers)
    
    try:
        vaults = glacier.list_vaults()
        
        for vault in vaults.get('VaultList', []):
            # Convert size to GB
            size_gb = vault.get('SizeInBytes', 0) / (1024**3)
            
            # Get notification configuration
            try:
                notification = glacier.get_vault_notifications(vaultName=vault['VaultName'])
                notifications_enabled = 'Yes'
                sns_topic = notification.get('vaultNotificationConfig', {}).get('SNSTopic', 'N/A')
            except:
                notifications_enabled = 'No'
                sns_topic = 'N/A'
            
            last_inventory = vault.get('LastInventoryDate')
            last_inventory_str = last_inventory.strftime('%Y-%m-%d %H:%M:%S') if last_inventory else 'N/A'
            
            ws.append([
                vault['VaultName'],
                vault['VaultARN'],
                vault['CreationDate'].strftime('%Y-%m-%d %H:%M:%S'),
                last_inventory_str,
                vault.get('NumberOfArchives', 0),
                f"{size_gb:.2f}",
                notifications_enabled,
                sns_topic
            ])
    except Exception as e:
        print(f"    Error: {e}")
    
    apply_header_style(ws, header_font, header_fill, header_alignment)

def export_cognito_user_pools(ws, cognito_idp, header_font, header_fill, header_alignment):
    """Export Cognito User Pools"""
    print("  - Exporting Cognito User Pools...")
    
    headers = [
        'User Pool Name', 'User Pool ID', 'Status', 'Creation Date',
        'Last Modified Date', 'MFA Configuration', 'Email Verification',
        'Phone Verification', 'Auto Verified Attributes', 'Password Policy',
        'Lambda Triggers', 'User Pool Clients', 'Domain Name',
        'Estimated Users'
    ]
    ws.append(headers)
    
    try:
        user_pools = cognito_idp.list_user_pools(MaxResults=60)
        
        for pool_summary in user_pools.get('UserPools', []):
            pool_id = pool_summary['Id']
            
            try:
                # Get detailed user pool information
                pool = cognito_idp.describe_user_pool(UserPoolId=pool_id)['UserPool']
                
                # MFA configuration
                mfa_config = pool.get('MfaConfiguration', 'OFF')
                
                # Email and phone verification
                auto_verified = ', '.join(pool.get('AutoVerifiedAttributes', [])) or 'None'
                
                # Password policy
                password_policy = pool.get('Policies', {}).get('PasswordPolicy', {})
                min_length = password_policy.get('MinimumLength', 'N/A')
                password_summary = f"Min Length: {min_length}"
                
                # Lambda triggers
                lambda_config = pool.get('LambdaConfig', {})
                lambda_triggers = len([k for k, v in lambda_config.items() if v]) if lambda_config else 0
                
                # Get user pool clients
                try:
                    clients = cognito_idp.list_user_pool_clients(UserPoolId=pool_id, MaxResults=60)
                    client_count = len(clients.get('UserPoolClients', []))
                except:
                    client_count = 0
                
                # Get domain name
                try:
                    domain = cognito_idp.describe_user_pool_domain(Domain=pool['Name'])
                    domain_name = domain.get('DomainDescription', {}).get('Domain', 'N/A')
                except:
                    domain_name = 'N/A'
                
                # Estimated number of users
                estimated_users = pool.get('EstimatedNumberOfUsers', 'N/A')
                
                ws.append([
                    pool['Name'],
                    pool['Id'],
                    pool.get('Status', 'N/A'),
                    pool['CreationDate'].strftime('%Y-%m-%d %H:%M:%S'),
                    pool['LastModifiedDate'].strftime('%Y-%m-%d %H:%M:%S'),
                    mfa_config,
                    'Yes' if 'email' in pool.get('AutoVerifiedAttributes', []) else 'No',
                    'Yes' if 'phone_number' in pool.get('AutoVerifiedAttributes', []) else 'No',
                    auto_verified,
                    password_summary,
                    lambda_triggers,
                    client_count,
                    domain_name,
                    estimated_users
                ])
            except Exception as e:
                print(f"    Error processing user pool {pool_id}: {e}")
    except Exception as e:
        print(f"    Error: {e}")
    
    apply_header_style(ws, header_font, header_fill, header_alignment)

def export_cognito_identity_pools(ws, cognito_identity, header_font, header_fill, header_alignment):
    """Export Cognito Identity Pools"""
    print("  - Exporting Cognito Identity Pools...")
    
    headers = [
        'Identity Pool Name', 'Identity Pool ID', 
        'Allow Unauthenticated Access', 'Identity Providers',
        'Cognito User Pools', 'SAML Providers', 
        'OpenID Connect Providers', 'Developer Providers'
    ]
    ws.append(headers)
    
    try:
        identity_pools = cognito_identity.list_identity_pools(MaxResults=60)
        
        for pool_summary in identity_pools.get('IdentityPools', []):
            pool_id = pool_summary['IdentityPoolId']
            
            try:
                # Get detailed identity pool information
                pool = cognito_identity.describe_identity_pool(IdentityPoolId=pool_id)
                
                # Count identity providers
                cognito_providers = pool.get('CognitoIdentityProviders', [])
                cognito_provider_count = len(cognito_providers)
                
                saml_providers = pool.get('SamlProviderARNs', [])
                saml_provider_count = len(saml_providers)
                
                oidc_providers = pool.get('OpenIdConnectProviderARNs', [])
                oidc_provider_count = len(oidc_providers)
                
                developer_providers = pool.get('DeveloperProviderName', 'None')
                
                # Supported login providers
                supported_providers = ', '.join(pool.get('SupportedLoginProviders', {}).keys()) or 'None'
                
                ws.append([
                    pool['IdentityPoolName'],
                    pool['IdentityPoolId'],
                    'Yes' if pool.get('AllowUnauthenticatedIdentities', False) else 'No',
                    supported_providers,
                    cognito_provider_count,
                    saml_provider_count,
                    oidc_provider_count,
                    developer_providers
                ])
            except Exception as e:
                print(f"    Error processing identity pool {pool_id}: {e}")
    except Exception as e:
        print(f"    Error: {e}")
    
    apply_header_style(ws, header_font, header_fill, header_alignment)

def export_aws_resources_for_profile(profile_name):
    """
    Export AWS resources for a specific profile to Excel.
    """
    print(f"\n{'='*70}")
    print(f"Processing AWS Profile: {profile_name}")
    print(f"{'='*70}")
    
    try:
        # Initialize AWS session with profile
        session = boto3.Session(profile_name=profile_name)
        
        # Get account information
        account_id, account_alias = get_account_info(session)
        region = session.region_name or 'us-east-1'
        
        print(f"Account ID: {account_id}")
        print(f"Account Alias: {account_alias}")
        print(f"Region: {region}")
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%y%m%d-%H%M')
        output_file = f'aws_resources_{timestamp}_{account_id}-{account_alias}.xlsx'
        
        # Create a new Excel workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Initialize AWS clients
        print("\n  - Initializing AWS clients...")
        ec2 = session.client('ec2')
        rds = session.client('rds')
        lambda_client = session.client('lambda')
        efs = session.client('efs')
        ecs = session.client('ecs')
        eks = session.client('eks')
        elasticache = session.client('elasticache')
        mq = session.client('mq')
        elbv2 = session.client('elbv2')
        personalize = session.client('personalize')
        cloudwatch = session.client('cloudwatch')
        logs = session.client('logs')
        dynamodb = session.client('dynamodb')
        transfer = session.client('transfer')
        s3 = session.client('s3')
        glacier = session.client('glacier')
        cognito_idp = session.client('cognito-idp')
        cognito_identity = session.client('cognito-identity')
        
        # Export all resources
        print("\nExporting resources:")
        
        # EC2 Instances
        ws_ec2 = wb.create_sheet("EC2 Instances")
        export_ec2_instances(ws_ec2, ec2, header_font, header_fill, header_alignment)
        
        # RDS Instances
        ws_rds = wb.create_sheet("RDS Instances")
        export_rds_instances(ws_rds, rds, ec2, header_font, header_fill, header_alignment)
        
        # RDS Clusters (Aurora) (NEW)
        ws_rds_clusters = wb.create_sheet("RDS Clusters")
        export_rds_clusters(ws_rds_clusters, rds, ec2, header_font, header_fill, header_alignment)
        
        # Lambda Functions
        ws_lambda = wb.create_sheet("Lambda Functions")
        export_lambda_functions(ws_lambda, lambda_client, ec2, header_font, header_fill, header_alignment)
        
        # EFS File Systems
        ws_efs = wb.create_sheet("EFS File Systems")
        export_efs_filesystems(ws_efs, efs, header_font, header_fill, header_alignment)
        
        # ECS Services
        ws_ecs = wb.create_sheet("ECS Services")
        export_ecs_services(ws_ecs, ecs, ec2, header_font, header_fill, header_alignment)
        
        # EKS Clusters
        ws_eks = wb.create_sheet("EKS Clusters")
        export_eks_clusters(ws_eks, eks, ec2, header_font, header_fill, header_alignment)
        
        # ElastiCache
        ws_elasticache = wb.create_sheet("ElastiCache")
        export_elasticache_clusters(ws_elasticache, elasticache, ec2, header_font, header_fill, header_alignment)
        
        # Amazon MQ
        ws_mq = wb.create_sheet("Amazon MQ")
        export_mq_brokers(ws_mq, mq, ec2, header_font, header_fill, header_alignment)
        
        # Load Balancers (NEW)
        ws_elb = wb.create_sheet("Load Balancers")
        export_load_balancers(ws_elb, elbv2, ec2, header_font, header_fill, header_alignment)
        
        # DynamoDB Tables (NEW)
        ws_dynamodb = wb.create_sheet("DynamoDB Tables")
        export_dynamodb_tables(ws_dynamodb, dynamodb, header_font, header_fill, header_alignment)
        
        # CloudWatch Alarms (NEW)
        ws_cw_alarms = wb.create_sheet("CloudWatch Alarms")
        export_cloudwatch_alarms(ws_cw_alarms, cloudwatch, header_font, header_fill, header_alignment)
        
        # CloudWatch Log Groups (NEW)
        ws_cw_logs = wb.create_sheet("CloudWatch Logs")
        export_cloudwatch_log_groups(ws_cw_logs, logs, header_font, header_fill, header_alignment)
        
        # Transfer Family (NEW)
        ws_transfer = wb.create_sheet("Transfer Family")
        export_transfer_family(ws_transfer, transfer, ec2, header_font, header_fill, header_alignment)
        
        # AWS Personalize (NEW)
        ws_personalize = wb.create_sheet("Personalize")
        export_personalize(ws_personalize, personalize, header_font, header_fill, header_alignment)

        # S3 Buckets
        ws_s3 = wb.create_sheet("S3 Buckets")
        export_s3_buckets(ws_s3, s3, header_font, header_fill, header_alignment)

        # S3 Glacier Vaults
        ws_glacier = wb.create_sheet("Glacier Vaults")
        export_s3_glacier_vaults(ws_glacier, glacier, header_font, header_fill, header_alignment)

        # Cognito User Pools
        ws_cognito_users = wb.create_sheet("Cognito User Pools")
        export_cognito_user_pools(ws_cognito_users, cognito_idp, header_font, header_fill, header_alignment)

        # Cognito Identity Pools
        ws_cognito_identity = wb.create_sheet("Cognito Identity Pools")
        export_cognito_identity_pools(ws_cognito_identity, cognito_identity, header_font, header_fill, header_alignment)
        
        # VPC Summary
        ws_vpc = wb.create_sheet("VPC Summary")
        export_vpc_summary(ws_vpc, ec2, region, header_font, header_fill, header_alignment)
        
        # Auto-adjust column widths for all sheets
        print("\n  - Adjusting column widths...")
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze top row for all sheets
        for sheet in wb.worksheets:
            sheet.freeze_panes = 'A2'
        
        # Save the workbook
        wb.save(output_file)
        print(f"\n✅ Export complete!")
        print(f"📊 File: {output_file}")
        print(f"📁 Sheets: {', '.join([ws.title for ws in wb.worksheets])}")
        
        return True
        
    except ProfileNotFound:
        print(f"❌ Error: AWS profile '{profile_name}' not found!")
        print(f"   Available profiles are configured in ~/.aws/credentials")
        return False
    except ClientError as e:
        print(f"❌ AWS Error: {e}")
        return False
    except Exception as e:
        print(f"❌ Unexpected Error: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """Main function to process multiple AWS profiles"""
    import sys
    
    # Use command line arguments if provided, otherwise use configured profiles
    if len(sys.argv) > 1:
        profiles_to_process = sys.argv[1:]
        print(f"Using profiles from command line: {', '.join(profiles_to_process)}")
    else:
        profiles_to_process = AWS_PROFILES
        print(f"Using profiles from configuration: {', '.join(profiles_to_process)}")
    
    print(f"\n{'='*70}")
    print(f"AWS Multi-Profile Resource Exporter")
    print(f"{'='*70}")
    print(f"Total profiles to process: {len(profiles_to_process)}")
    
    # Track results
    successful = []
    failed = []
    
    # Process each profile
    for profile in profiles_to_process:
        if export_aws_resources_for_profile(profile):
            successful.append(profile)
        else:
            failed.append(profile)
    
    # Summary
    print(f"\n{'='*70}")
    print(f"SUMMARY")
    print(f"{'='*70}")
    print(f"✅ Successfully processed: {len(successful)} profile(s)")
    if successful:
        for profile in successful:
            print(f"   - {profile}")
    
    if failed:
        print(f"\n❌ Failed to process: {len(failed)} profile(s)")
        for profile in failed:
            print(f"   - {profile}")
    
    print(f"\n{'='*70}")
    print(f"All exports completed!")
    print(f"{'='*70}\n")

if __name__ == "__main__":
    # Make sure AWS credentials are configured for each profile
    # Configure profiles in ~/.aws/credentials or ~/.aws/config
    # 
    # Example ~/.aws/credentials:
    # [default]
    # aws_access_key_id = YOUR_KEY
    # aws_secret_access_key = YOUR_SECRET
    # 
    # [production]
    # aws_access_key_id = YOUR_KEY
    # aws_secret_access_key = YOUR_SECRET
    # region = us-west-2
    # 
    # [staging]
    # aws_access_key_id = YOUR_KEY
    # aws_secret_access_key = YOUR_SECRET
    # region = us-east-1
    
    main()